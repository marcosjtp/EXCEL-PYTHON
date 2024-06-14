import tkinter
#from tkinter import *
from tkinter import  filedialog ,Tk,Entry,StringVar
import openpyxl

#from fpdf.fpdf import FPDF
#import subprocess
import csv
from tkinter import messagebox
import os
import subprocess    
varia1=""
wre = ""
peso = ""
cliente=""
desc=""
cobrar=""
listas=[["WR","PESO"]]
lista=[]
hoja=""
root=tkinter.Tk(screenName="Convertidor", baseName=None, className='Tix')
root.geometry("370x300+500+200")
root.resizable(0,0)
root.wm_attributes("-topmost", 1)
#root.iconbitmap(default='Libros.ico')
Entry1=Entry(root)
Entry1.place(rely=0.03,relx=0.045,relheight=0.1,relwidth=0.90)
Entry1.config(background="grey",textvariable=varia1)

Entry2=Entry(root)
Entry2.place(rely=0.5,relx=0.28,relheight=0.1,relwidth=0.40)
Entry2.config(background="grey",foreground="red")

def importarExcel():
    try:
        
    

        user = os.environ.get("USERNAME")
        iniciardialogo = "C:\\Users\\" + user + "\\Desktop"
        File = filedialog.askopenfilenames(
            title="Abrir Archivos",
            initialdir=iniciardialogo,
            filetypes=[
                ("Excel", "*.xlsm"),
                ("Excel", "*.xltx"),
                ("Excel", "*.xltm"),
                ("Excel", "*.xlsx"),
                ("Excel", "*.csv"),
            ],
        )
        # print(File[0])
        path = File[0]
        
        Entry1.delete(0, tkinter.END)
        Entry1.insert(0,path)
    except:
        messagebox.showerror("ERROR","ERROR AL CARGAR EL ARCHIVO")
        next
def limpiar():
    listas.clear()
    lista.clear()
    
    print(lista)
    print(listas)

def exportarcsv():
    with open(Entry1.get(), newline='') as csvfile:
        reader = csv.reader(csvfile, delimiter=',')
        if os.path.isfile("Factura.csv")==True:
            os.startfile(".\\Factura.csv")
            for row in reader:
                print(', '.join(row))
def exportarexcel():
    try:     
    
        workbook = openpyxl.load_workbook(Entry1.get())
        #print(path)
        print(Entry1.get())
        sheet = workbook.active
        contador = 0

        list_values = list(sheet.values)
        for row in list_values[1:]:
            contador += 1
            lista=[]
            wre = row[0]
            #cliente=row[1]
            #desc=row[2]
            peso = row[1]
            #cobrar=row[4]
            #cliente=cliente.replace("LP-31-D1-","").replace(" ","")
            peso = (
                peso.replace("lbs", "")
                .replace(" 0 oz", ".0")
                .replace(" 1 oz", ".0625")
                .replace(" 2 oz", ".125")
                .replace(" 3 oz", ".1875")
                .replace(" 4 oz", ".25")
                .replace(" 5 oz", ".3125")
                .replace(" 6 oz", ".375")
                .replace(" 7 oz", ".4375")
                .replace(" 8 oz", ".5")
                .replace(" 9 oz", ".5625")
                .replace(" 10 oz", ".625")
                .replace(" 11 oz", ".6875")
                .replace(" 12 oz", ".75")
                .replace(" 13 oz", ".8125")
                .replace(" 14 oz", ".875")
                .replace(" 15 oz", ".9375")
                .replace(" ", "")
                
            )
            #cobrar=(cobrar.replace("RD$","").replace(",",""))
            wre=wre.replace("LP-0","LP-")
            lista.append(wre)
            #lista.append(cliente)
            #lista.append(desc)
            lista.append(float(peso))
            #lista.append(float(cobrar))
            listas.append(lista)
            print(lista)
            
        # Crear un nuevo libro de trabajo (workbook)
        wb = openpyxl.Workbook()
        # Obtener la hoja activa
        
        wb.create_sheet(Entry2.get())
        sheet = wb[Entry2.get()]  
        # Escribir los datos en la hoja de cálculo
        for fila in listas:
            sheet.append(fila)   
        # Guardar el archivo
        wb.save(str(Entry2.get())+".xlsx")
        limpiar()

        if os.path.isfile(str(Entry2.get())+".xlsx")==True:
            messagebox.showinfo("MENSAJE","SE HA MODIFICADO ARCHIVO CORRECTAMENTE")
            os.startfile(".\\"+str(Entry2.get())+".xlsx")
            print(listas)
            
    except:
        messagebox.showerror("ERROR","ERROR AL MODIFICAR EL ARCHIVO")
        next

btnreiniciar=tkinter.Button(root,text="ABRIR EXCEL",background="gold",command=importarExcel)
btnreiniciar.place( x=65, y = 70)
btnreiniciar.config(width=30, height=2)
btnactivar=tkinter.Button(root,text="EXPORTAR EXCEL",background="lime green",command=exportarexcel)
btnactivar.place( x=65, y = 200)
btnactivar.config(width=30, height=2) 
root.mainloop()





